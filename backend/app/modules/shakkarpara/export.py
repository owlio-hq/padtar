"""Excel + PDF export for Shakkarpara batches.

Each batch is rendered as its own formatted block (date, ingredient table with
rate/usage/total colored to match the app + original Excel, production/total/
padtar summary) — never a flat data dump. Notes get their own clearly separate
section: a dedicated "Notes" sheet in the Excel workbook, and a dedicated
section at the end of the PDF — never mixed into the data tables.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

from app.core.export_style import (
    BORDER_COLOR,
    fit_to_one_page,
    HEADER_FILL,
    OIL_SIT_FILL,
    OIL_SIT_TEXT,
    PADTAR_FILL,
    PADTAR_TEXT,
    RATE_FILL,
    RATE_TEXT,
    SUBTOTAL_FILL,
    SUBTOTAL_TEXT,
    TOTAL_FILL,
    TOTAL_TEXT,
    USAGE_FILL,
    USAGE_TEXT,
)
from app.core.notes import parse_notes
from app.core.pdf import BODY_STYLE, SECTION_STYLE, new_document, notes_section, spacer, title_block
from app.modules.shakkarpara.schemas import BatchOut

_thin_side = Side(style="thin", color=BORDER_COLOR)
THIN_BORDER = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def build_excel(batches: list[BatchOut]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Batches"

    col_widths = [22, 12, 12, 14, 14]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    row = 1
    for batch in sorted(batches, key=lambda b: b.date):
        ws.cell(row=row, column=1, value="Date").font = Font(bold=True)
        ws.cell(row=row, column=2, value=batch.date.strftime("%d %b %Y")).font = Font(bold=True)
        row += 1

        headers = ["Ingredient", "Rate (₹)", "Usage", "Unit", "Total (₹)"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = _fill(HEADER_FILL)
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
        row += 1

        for ing in batch.ingredients:
            ws.cell(row=row, column=1, value=ing.name).border = THIN_BORDER
            rate_cell = ws.cell(row=row, column=2, value=ing.rate)
            rate_cell.fill, rate_cell.font, rate_cell.border = _fill(RATE_FILL), Font(color=RATE_TEXT), THIN_BORDER
            usage_cell = ws.cell(row=row, column=3, value=round(ing.usage, 4))
            usage_cell.fill, usage_cell.font, usage_cell.border = _fill(USAGE_FILL), Font(color=USAGE_TEXT), THIN_BORDER
            ws.cell(row=row, column=4, value=ing.unit).border = THIN_BORDER
            total_cell = ws.cell(row=row, column=5, value=round(ing.total, 2))
            total_cell.fill, total_cell.font, total_cell.border = _fill(TOTAL_FILL), Font(color=TOTAL_TEXT, bold=True), THIN_BORDER
            row += 1

        ws.cell(row=row, column=1, value="Production").font = Font(italic=True)
        ws.cell(row=row, column=2, value=f"{batch.production_qty} {batch.production_unit}")
        row += 1

        # Batch Total row — stronger green fill to stand out from ingredient totals column.
        total_label = ws.cell(row=row, column=1, value="Total")
        total_label.fill, total_label.font = _fill(SUBTOTAL_FILL), Font(color=SUBTOTAL_TEXT, bold=True)
        total_cell = ws.cell(row=row, column=5, value=round(batch.total, 2))
        total_cell.fill, total_cell.font = _fill(SUBTOTAL_FILL), Font(color=SUBTOTAL_TEXT, bold=True)
        row += 1

        if batch.extra_per_unit:
            ws.cell(row=row, column=1, value="Office Expenses").font = Font(italic=True)
            ws.cell(row=row, column=2, value=round(batch.extra_per_unit, 2))
            row += 1

        # Padtar row — warm amber, distinct final-cost highlight.
        padtar_label = ws.cell(row=row, column=1, value="Padtar")
        padtar_label.fill, padtar_label.font = _fill(PADTAR_FILL), Font(color=PADTAR_TEXT, bold=True)
        padtar_cell = ws.cell(row=row, column=2, value=round(batch.padtar, 2) if batch.padtar is not None else None)
        padtar_cell.fill, padtar_cell.font = _fill(PADTAR_FILL), Font(color=PADTAR_TEXT, bold=True)
        row += 1

        # Oil-sit sub-table — nava + juna + toppa − parat = net vaprash (dabba).
        if batch.oil_sit is not None:
            row += 1  # small gap
            oil_header = ws.cell(row=row, column=1, value="Oil Sheet")
            oil_header.fill, oil_header.font = _fill(OIL_SIT_FILL), Font(color=OIL_SIT_TEXT, bold=True)
            for c in range(2, 6):
                ws.cell(row=row, column=c).fill = _fill(OIL_SIT_FILL)
            row += 1
            oil_headers = ["Nava dabba", "Juna dabba", "Toppa", "Parat malela", "Net vaprash"]
            for col, h in enumerate(oil_headers, start=1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.fill, cell.font, cell.border = _fill(HEADER_FILL), Font(bold=True), THIN_BORDER
            row += 1
            oil = batch.oil_sit
            for col, v in enumerate(
                [oil.nava_dabba, oil.juna_dabba, oil.toppa, oil.parat_malela, oil.net_vaprash], start=1
            ):
                cell = ws.cell(row=row, column=col, value=round(v, 4) if v is not None else None)
                cell.border = THIN_BORDER
                if col == 5:  # net vaprash — highlight
                    cell.fill, cell.font = _fill(OIL_SIT_FILL), Font(color=OIL_SIT_TEXT, bold=True)
            row += 1

        row += 1  # blank separator row before the next batch block

    notes_ws = wb.create_sheet("Notes")
    notes_ws.column_dimensions["A"].width = 16
    notes_ws.column_dimensions["B"].width = 45
    notes_ws.column_dimensions["C"].width = 35
    notes_ws.append(["Date", "Note", "Detail"])
    for cell in notes_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _fill(HEADER_FILL)
    for batch in sorted(batches, key=lambda b: b.date):
        for i, (note, detail) in enumerate(parse_notes(batch.notes)):
            notes_ws.append([batch.date.strftime("%d %b %Y") if i == 0 else "", note, detail])
            for cell in notes_ws[notes_ws.max_row]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    fit_to_one_page(ws)
    fit_to_one_page(notes_ws)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_pdf(batches: list[BatchOut]) -> bytes:
    buffer = BytesIO()
    doc = new_document(buffer, "Shakkarpara")
    story = title_block("Shakkarpara", "Batch costing export")

    for batch in sorted(batches, key=lambda b: b.date):
        story.append(Paragraph(f"Batch — {batch.date.strftime('%d %b %Y')}", SECTION_STYLE))

        # PDF uses Helvetica, which has no ₹ glyph (renders as a black box) — use "Rs."
        rows = [["Ingredient", "Rate (Rs.)", "Usage", "Unit", "Total (Rs.)"]]
        for ing in batch.ingredients:
            rows.append([ing.name, f"{ing.rate:g}", f"{ing.usage:.2f}", ing.unit, f"{ing.total:.2f}"])
        padtar_label = "Padtar"
        rows.append(["Production", "", "", "", f"{batch.production_qty} {batch.production_unit}"])
        rows.append(["Total", "", "", "", f"{batch.total:.2f}"])
        rows.append([padtar_label, "", "", "", f"{batch.padtar:.2f}" if batch.padtar is not None else "—"])

        table = Table(rows, colWidths=[5 * cm, 2.5 * cm, 2.5 * cm, 3 * cm, 3 * cm])
        n = len(batch.ingredients)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FILL}")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (1, 1), (1, n), colors.HexColor(f"#{RATE_FILL}")),
            ("BACKGROUND", (2, 1), (2, n), colors.HexColor(f"#{USAGE_FILL}")),
            ("BACKGROUND", (4, 1), (4, n), colors.HexColor(f"#{TOTAL_FILL}")),
            # Batch Total row — stronger green.
            ("BACKGROUND", (0, n + 2), (-1, n + 2), colors.HexColor(f"#{SUBTOTAL_FILL}")),
            ("TEXTCOLOR", (0, n + 2), (-1, n + 2), colors.HexColor(f"#{SUBTOTAL_TEXT}")),
            ("FONTNAME", (0, n + 2), (-1, n + 2), "Helvetica-Bold"),
            # Padtar row — warm amber.
            ("BACKGROUND", (0, n + 3), (-1, n + 3), colors.HexColor(f"#{PADTAR_FILL}")),
            ("TEXTCOLOR", (0, n + 3), (-1, n + 3), colors.HexColor(f"#{PADTAR_TEXT}")),
            ("FONTNAME", (0, n + 3), (-1, n + 3), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{BORDER_COLOR}")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ]
        table.setStyle(TableStyle(style))
        story.append(table)

        # Oil-sit sub-table (below the main ingredient table).
        if batch.oil_sit is not None:
            story.append(spacer(0.15))
            oil = batch.oil_sit
            oil_rows = [
                ["Oil Sheet", "", "", "", ""],
                ["Nava dabba", "Juna dabba", "Toppa", "Parat malela", "Net vaprash"],
                [
                    f"{oil.nava_dabba:g}",
                    f"{oil.juna_dabba:g}",
                    f"{oil.toppa:g}",
                    f"{oil.parat_malela:g}",
                    f"{oil.net_vaprash:g}" if oil.net_vaprash is not None else "—",
                ],
            ]
            oil_table = Table(oil_rows, colWidths=[3.2 * cm, 3.2 * cm, 3.2 * cm, 3.2 * cm, 3.2 * cm])
            oil_table.setStyle(TableStyle([
                ("SPAN", (0, 0), (-1, 0)),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{OIL_SIT_FILL}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(f"#{OIL_SIT_TEXT}")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor(f"#{HEADER_FILL}")),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("BACKGROUND", (4, 2), (4, 2), colors.HexColor(f"#{OIL_SIT_FILL}")),
                ("TEXTCOLOR", (4, 2), (4, 2), colors.HexColor(f"#{OIL_SIT_TEXT}")),
                ("FONTNAME", (4, 2), (4, 2), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{BORDER_COLOR}")),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(oil_table)

        story.append(spacer(0.35))

    note_entries = [(b.date.strftime("%d %b %Y"), parse_notes(b.notes)) for b in sorted(batches, key=lambda b: b.date) if b.notes]
    story.extend(notes_section(note_entries))

    if not batches:
        story.append(Paragraph("No batches to export.", BODY_STYLE))

    doc.build(story)
    return buffer.getvalue()
