"""Excel + PDF export for Rojmel — two independent exports matching the two
pages in the UI: day entries (sales + cashbook) and monthly stock. Notes get
their own clearly separate section in both formats, same rule as Shakkarpara.
"""

import calendar
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Table, TableStyle

from app.core.export_style import (
    BORDER_COLOR,
    fit_to_one_page,
    HEADER_FILL,
    NEGATIVE_FILL,
    NEGATIVE_TEXT,
    PADTAR_FILL,
    PADTAR_TEXT,
    RATE_FILL,
    RATE_TEXT,
    SUBTOTAL_FILL,
    SUBTOTAL_TEXT,
    TOTAL_FILL,
    TOTAL_TEXT,
    USAGE_FILL,
    USAGE_TEXT,
)
from app.core.notes import parse_notes
from app.core.pdf import BODY_STYLE, SECTION_STYLE, new_document, notes_section, spacer, title_block
from app.modules.rojmel.schemas import DayOut, StockRowOut

_thin_side = Side(style="thin", color=BORDER_COLOR)
_bold_side = Side(style="medium", color="000000")
THIN_BORDER = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)


def _bold_outline(ws, top_row: int, bottom_row: int, left_col: int, right_col: int) -> None:
    """Apply a bold (medium) black outer border around a rectangular cell range."""
    for r in range(top_row, bottom_row + 1):
        for c in range(left_col, right_col + 1):
            cell = ws.cell(row=r, column=c)
            existing = cell.border
            new_left = _bold_side if c == left_col else existing.left
            new_right = _bold_side if c == right_col else existing.right
            new_top = _bold_side if r == top_row else existing.top
            new_bottom = _bold_side if r == bottom_row else existing.bottom
            cell.border = Border(left=new_left, right=new_right, top=new_top, bottom=new_bottom)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _money_table(title: str, lines, head_fill: str, head_text: str) -> Table:
    """One PDF money table (Income or Kharcho) with a coloured header row.
    Columns are Amount | Description | Note, matching the on-screen order; the
    block name sits over the wide description column."""
    rows = [["Amount", title, "Note"]]
    for m in lines:
        rows.append([f"{m.amount:.2f}", m.description, m.note])
    if len(rows) == 1:
        rows.append(["", "—", ""])
    tbl = Table(rows, colWidths=[2.0 * cm, 4.6 * cm, 2.6 * cm])
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{head_fill}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(f"#{head_text}")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{BORDER_COLOR}")),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (0, -1), "RIGHT"),  # amounts line up
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ]
        )
    )
    return tbl


def build_days_excel(days: list[DayOut]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Days"
    # widths: Product | Rate | OPP.PIC | CLO.PIC | NET.PIC | Sales | Total
    # These also serve the side-by-side money layout (A-C Income, E-G Kharcho).
    for idx, width in enumerate([22, 12, 14, 11, 18, 12, 14], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    row = 1
    for day in sorted(days, key=lambda d: d.date):
        ws.cell(row=row, column=1, value="Date").font = Font(bold=True)
        ws.cell(row=row, column=2, value=day.date.strftime("%d %b %Y")).font = Font(bold=True)
        row += 1

        # Column order: Product | Rate | OPP.PIC | CLO.PIC | NET.PIC | Sales | Total
        # (the stock trio sits together, then the count they type, then the money)
        sales_header_row = row
        for col, header in enumerate(["Product", "Rate (₹)", "OPP.PIC", "CLO.PIC", "NET.PIC", "Sales", "Total (₹)"], start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill, cell.font, cell.border = _fill(HEADER_FILL), Font(bold=True), THIN_BORDER
        row += 1

        right, center = Alignment(horizontal="right"), Alignment(horizontal="center")
        for s in day.sales_lines:
            ws.cell(row=row, column=1, value=s.product).border = THIN_BORDER
            c = ws.cell(row=row, column=2, value=s.rate)
            c.fill, c.font, c.border, c.alignment = _fill(RATE_FILL), Font(color=RATE_TEXT), THIN_BORDER, center
            for col, val in ((3, s.opening_pic), (4, s.closing_pic)):
                c = ws.cell(row=row, column=col, value=val)
                c.border, c.alignment = THIN_BORDER, right
            net_cell = ws.cell(row=row, column=5, value=s.net_pic)
            fill, text = (NEGATIVE_FILL, NEGATIVE_TEXT) if s.net_pic < 0 else (TOTAL_FILL, TOTAL_TEXT)
            net_cell.fill, net_cell.font, net_cell.border = _fill(fill), Font(color=text, bold=True), THIN_BORDER
            net_cell.alignment = right
            c = ws.cell(row=row, column=6, value=s.qty)
            c.fill, c.font, c.border, c.alignment = _fill(USAGE_FILL), Font(color=USAGE_TEXT), THIN_BORDER, right
            c = ws.cell(row=row, column=7, value=round(s.total, 2))
            c.fill, c.font, c.border, c.alignment = _fill(TOTAL_FILL), Font(color=TOTAL_TEXT, bold=True), THIN_BORDER, right
            row += 1

        fs_label = ws.cell(row=row, column=1, value="Factory Sales")
        fs_label.fill, fs_label.font = _fill(SUBTOTAL_FILL), Font(color=SUBTOTAL_TEXT, bold=True)
        c = ws.cell(row=row, column=7, value=round(day.factory_sales, 2))
        c.fill, c.font = _fill(SUBTOTAL_FILL), Font(color=SUBTOTAL_TEXT, bold=True)
        _bold_outline(ws, sales_header_row, row, 1, 7)
        row += 2

        # Income (cols A-C) and Kharcho (cols E-G) side-by-side.
        # Column order: Description | Note | Amount (₹) — amount on the right.
        # Thick red right-border on col C separates the two sides visually.
        _red_side = Side(style="medium", color="CC0000")
        income_lines = day.income_lines or []
        expense_lines = day.expense_lines or []

        # Headers
        for label, start_col in (("Income", 1), ("Kharcho", 5)):
            ws.cell(row=row, column=start_col, value=label).font = Font(bold=True, italic=True)
        row += 1
        for start_col in (1, 5):
            for offset, header in enumerate(["Description", "Note", "Amount (₹)"]):
                cell = ws.cell(row=row, column=start_col + offset, value=header)
                cell.fill, cell.font = _fill(HEADER_FILL), Font(bold=True)
                if offset == 2:
                    cell.alignment = Alignment(horizontal="right")
        money_header_row = row
        row += 1

        max_lines = max(len(income_lines), len(expense_lines), 1)
        for i in range(max_lines):
            if i < len(income_lines):
                m = income_lines[i]
                ws.cell(row=row, column=1, value=m.description)
                ws.cell(row=row, column=2, value=m.note)
                amt = ws.cell(row=row, column=3, value=m.amount)
                amt.alignment = Alignment(horizontal="right")
            if i < len(expense_lines):
                m = expense_lines[i]
                ws.cell(row=row, column=5, value=m.description)
                ws.cell(row=row, column=6, value=m.note)
                amt = ws.cell(row=row, column=7, value=m.amount)
                amt.alignment = Alignment(horizontal="right")
            row += 1

        # Light borders on money cells
        for r in range(money_header_row, row):
            for c in (1, 2, 3, 5, 6, 7):
                ws.cell(row=r, column=c).border = THIN_BORDER

        # Bold outlines on Income (A-C) and Kharcho (E-G)
        _bold_outline(ws, money_header_row, row - 1, 1, 3)
        _bold_outline(ws, money_header_row, row - 1, 5, 7)
        # Red right border on col C — applied LAST so it's not overwritten
        for r in range(money_header_row, row):
            c3 = ws.cell(row=r, column=3)
            c3.border = Border(
                left=c3.border.left, top=c3.border.top, bottom=c3.border.bottom,
                right=_red_side,
            )

        row += 1

        # Carry Forward (cols A-C) and Notes (cols E-G) side-by-side — both 3 columns
        parsed_notes = parse_notes(day.notes) if day.notes else []
        has_cf = bool(day.carry_forward_lines)
        has_notes = bool(parsed_notes) and parsed_notes != [("", "")]

        if has_cf or has_notes:
            if has_cf:
                ws.cell(row=row, column=1, value="Carry Forward").font = Font(bold=True, italic=True)
            if has_notes:
                ws.cell(row=row, column=5, value="Notes").font = Font(bold=True, italic=True)
            row += 1

            cf_header_row = row
            if has_cf:
                cell = ws.cell(row=row, column=1, value="Name")
                cell.fill, cell.font, cell.border = _fill(HEADER_FILL), Font(bold=True), THIN_BORDER
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                cell = ws.cell(row=row, column=3, value="Carry Forward (₹)")
                cell.fill, cell.font, cell.border = _fill(HEADER_FILL), Font(bold=True), THIN_BORDER
                cell.alignment = Alignment(horizontal="right")
            if has_notes:
                for offset, header in enumerate(["Date", "Note", "Amount (₹)"]):
                    cell = ws.cell(row=row, column=5 + offset, value=header)
                    cell.fill, cell.font, cell.border = _fill(HEADER_FILL), Font(bold=True), THIN_BORDER
                ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
            row += 1

            cf_lines = day.carry_forward_lines or []
            max_side = max(len(cf_lines) + (1 if has_cf else 0), len(parsed_notes))
            for i in range(max_side):
                if has_cf and i < len(cf_lines):
                    ws.cell(row=row, column=1, value=cf_lines[i].name)
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                    c = ws.cell(row=row, column=3, value=cf_lines[i].amount)
                    c.border, c.alignment = THIN_BORDER, Alignment(horizontal="right")
                elif has_cf and i == len(cf_lines):
                    ws.cell(row=row, column=1, value="Total")
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                    ws.cell(row=row, column=1).fill = _fill(SUBTOTAL_FILL)
                    ws.cell(row=row, column=1).font = Font(color=SUBTOTAL_TEXT, bold=True)
                    c = ws.cell(row=row, column=3, value=round(sum(cf.amount for cf in cf_lines), 2))
                    c.fill, c.font, c.border = _fill(SUBTOTAL_FILL), Font(color=SUBTOTAL_TEXT, bold=True), THIN_BORDER
                    c.alignment = Alignment(horizontal="right")
                if has_notes and i < len(parsed_notes):
                    note_text, detail = parsed_notes[i]
                    ws.cell(row=row, column=5, value=day.date.strftime("%d %b %Y") if i == 0 else "").border = THIN_BORDER
                    ws.cell(row=row, column=6, value=note_text).border = THIN_BORDER
                    c = ws.cell(row=row, column=7, value=detail)
                    c.border, c.alignment = THIN_BORDER, Alignment(horizontal="right")
                row += 1

            if has_cf:
                _bold_outline(ws, cf_header_row, row - 1, 1, 3)
            if has_notes:
                _bold_outline(ws, cf_header_row, row - 1, 5, 7)

            row += 1

        # Summary — stacked vertically: Income, Kharcho, Cash on Hand (below carry forward/notes)
        summary_start = row
        for label, val, fill_c, text_c in (
            ("Income:", round(day.total_income, 2), SUBTOTAL_FILL, SUBTOTAL_TEXT),
            ("Kharcho:", round(day.total_expense, 2), NEGATIVE_FILL, NEGATIVE_TEXT),
            ("Cash on Hand:", round(day.cash_on_hand, 2), PADTAR_FILL, PADTAR_TEXT),
        ):
            lbl = ws.cell(row=row, column=1, value=label)
            lbl.font = Font(bold=True, color=text_c)
            lbl.fill = _fill(fill_c)
            lbl.border = THIN_BORDER
            v = ws.cell(row=row, column=2, value=val)
            v.font = Font(bold=True, color=text_c)
            v.fill = _fill(fill_c)
            v.alignment = Alignment(horizontal="right")
            v.border = THIN_BORDER
            row += 1
        _bold_outline(ws, summary_start, row - 1, 1, 2)
        row += 2

    fit_to_one_page(ws)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_days_pdf(days: list[DayOut]) -> bytes:
    buffer = BytesIO()
    # visible word is "Rojmed"; module/table/file names stay "rojmel"
    doc = new_document(buffer, "Rojmed")
    story = title_block("Rojmed", "Daily sales & cash export")

    for day in sorted(days, key=lambda d: d.date):
        story.append(Paragraph(f"Day — {day.date.strftime('%d %b %Y')}", SECTION_STYLE))

        # PDF uses Helvetica, which has no ₹ glyph (renders as a black box) — use "Rs."
        # Column order: Product | Rate | OPP.PIC | CLO.PIC | NET.PIC | Sales | Total
        rows = [["Product", "Rate (Rs.)", "OPP.PIC", "CLO.PIC", "NET.PIC", "Sales", "Total (Rs.)"]]
        for s in day.sales_lines:
            rows.append([s.product, f"{s.rate:g}", f"{s.opening_pic:g}", f"{s.closing_pic:g}", f"{s.net_pic:g}", f"{s.qty:g}", f"{s.total:.2f}"])
        rows.append(["Factory Sales", "", "", "", "", "", f"{day.factory_sales:.2f}"])
        n = len(day.sales_lines)
        table = Table(rows, colWidths=[5.2 * cm, 2.2 * cm, 2 * cm, 2 * cm, 2 * cm, 1.8 * cm, 3.2 * cm])
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FILL}")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (1, 1), (1, n), colors.HexColor(f"#{RATE_FILL}")),
            ("BACKGROUND", (5, 1), (5, n), colors.HexColor(f"#{USAGE_FILL}")),
            ("BACKGROUND", (6, 1), (6, n), colors.HexColor(f"#{TOTAL_FILL}")),
            # Factory Sales row — stronger green.
            ("BACKGROUND", (0, n + 1), (-1, n + 1), colors.HexColor(f"#{SUBTOTAL_FILL}")),
            ("TEXTCOLOR", (0, n + 1), (-1, n + 1), colors.HexColor(f"#{SUBTOTAL_TEXT}")),
            ("FONTNAME", (0, n + 1), (-1, n + 1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{BORDER_COLOR}")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            # amounts/counts right, rate centred, product name left
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ]
        # NET.PIC cell red when negative, green otherwise (col index 4 after the reorder).
        for i, s in enumerate(day.sales_lines, start=1):
            fill, text = (NEGATIVE_FILL, NEGATIVE_TEXT) if s.net_pic < 0 else (TOTAL_FILL, TOTAL_TEXT)
            style.append(("BACKGROUND", (4, i), (4, i), colors.HexColor(f"#{fill}")))
            style.append(("TEXTCOLOR", (4, i), (4, i), colors.HexColor(f"#{text}")))
        table.setStyle(TableStyle(style))
        story.append(table)
        story.append(spacer(0.2))

        # Two separate money tables, side by side, matching the UI: Income (green
        # header) on the left, Kharcho (red header) on the right.
        income_tbl = _money_table("Income", day.income_lines, SUBTOTAL_FILL, SUBTOTAL_TEXT)
        expense_tbl = _money_table("Kharcho", day.expense_lines, NEGATIVE_FILL, NEGATIVE_TEXT)
        pair = Table([[income_tbl, expense_tbl]], colWidths=[9.2 * cm, 9.2 * cm])
        pair.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (0, 0), 6)]))
        story.append(pair)
        story.append(spacer(0.2))

        summary = Table(
            [["Total Income", f"{day.total_income:.2f}", "Total Expense", f"{day.total_expense:.2f}", "Cash on Hand", f"{day.cash_on_hand:.2f}"]],
            colWidths=[3.2 * cm, 2.7 * cm, 3.2 * cm, 2.7 * cm, 3.2 * cm, 3.4 * cm],
        )
        summary.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (1, 0), colors.HexColor(f"#{SUBTOTAL_FILL}")),
                    ("TEXTCOLOR", (0, 0), (1, 0), colors.HexColor(f"#{SUBTOTAL_TEXT}")),
                    ("BACKGROUND", (2, 0), (3, 0), colors.HexColor(f"#{NEGATIVE_FILL}")),
                    ("TEXTCOLOR", (2, 0), (3, 0), colors.HexColor(f"#{NEGATIVE_TEXT}")),
                    ("BACKGROUND", (4, 0), (5, 0), colors.HexColor(f"#{PADTAR_FILL}")),
                    ("TEXTCOLOR", (4, 0), (5, 0), colors.HexColor(f"#{PADTAR_TEXT}")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{BORDER_COLOR}")),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ]
            )
        )
        story.append(summary)
        story.append(spacer(0.2))

        if day.carry_forward_lines:
            cf_rows = [["Carry Forward", "Amount (Rs.)"]]
            for cf in day.carry_forward_lines:
                cf_rows.append([cf.name, f"{cf.amount:g}"])
            # Informational total — deliberately NOT part of Cash on Hand.
            cf_rows.append(["Total", f"{sum(cf.amount for cf in day.carry_forward_lines):.2f}"])
            cf_last = len(cf_rows) - 1
            cf_table = Table(cf_rows, colWidths=[8 * cm, 4 * cm], hAlign='LEFT')
            cf_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FILL}")),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("BACKGROUND", (0, cf_last), (-1, cf_last), colors.HexColor(f"#{SUBTOTAL_FILL}")),
                        ("TEXTCOLOR", (0, cf_last), (-1, cf_last), colors.HexColor(f"#{SUBTOTAL_TEXT}")),
                        ("FONTNAME", (0, cf_last), (-1, cf_last), "Helvetica-Bold"),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{BORDER_COLOR}")),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                    ]
                )
            )
            story.append(cf_table)
        story.append(spacer(0.4))

    note_entries = [(d.date.strftime("%d %b %Y"), parse_notes(d.notes)) for d in sorted(days, key=lambda d: d.date) if d.notes]
    story.extend(notes_section(note_entries))

    if not days:
        story.append(Paragraph("No days to export.", BODY_STYLE))

    doc.build(story)
    return buffer.getvalue()


def build_stock_excel(rows: list[StockRowOut], year: int, month: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append([f"{calendar.month_name[month]} {year}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    headers = ["Product", "Rate", "OPP.PIC (Opening)", "CLO.PIC (Closing)", "NET.PIC (Net)"]
    ws.append(headers)
    for cell in ws[3]:
        cell.font, cell.fill = Font(bold=True), _fill(HEADER_FILL)

    for row in rows:
        ws.append([row.product, row.rate, row.opening_pic, row.closing_pic, row.net_pic])
        net_cell = ws.cell(row=ws.max_row, column=5)
        if row.net_pic < 0:
            net_cell.fill, net_cell.font = _fill(NEGATIVE_FILL), Font(color=NEGATIVE_TEXT, bold=True)
        else:
            net_cell.fill, net_cell.font = _fill(PADTAR_FILL), Font(color=PADTAR_TEXT, bold=True)

    for idx, width in enumerate([20, 12, 16, 16, 14], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    fit_to_one_page(ws)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_stock_pdf(rows: list[StockRowOut], year: int, month: int) -> bytes:
    buffer = BytesIO()
    doc = new_document(buffer, "Rojmed Stock")
    story = title_block("Rojmed — Stock", f"{calendar.month_name[month]} {year}")

    table_rows = [["Product", "Rate", "Opening", "Closing", "Net"]]
    for row in rows:
        table_rows.append([row.product, f"{row.rate:g}", f"{row.opening_pic:g}", f"{row.closing_pic:g}", f"{row.net_pic:g}"])

    table = Table(table_rows, colWidths=[5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FILL}")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{BORDER_COLOR}")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]
    for i, row in enumerate(rows, start=1):
        fill = NEGATIVE_FILL if row.net_pic < 0 else PADTAR_FILL
        style.append(("BACKGROUND", (4, i), (4, i), colors.HexColor(f"#{fill}")))
    table.setStyle(TableStyle(style))
    story.append(table)

    doc.build(story)
    return buffer.getvalue()
