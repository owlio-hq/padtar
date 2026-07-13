"""Import selected real historical data into the app database.

Careful + verified: for every Shakkarpara batch we recompute total & padtar with
the app's own engine and assert they EXACTLY match the Excel's cached values
before saving (same cell-for-cell standard as the test suite). Anything that
fails verification, or is incomplete (no production), is SKIPPED and reported —
never silently imported wrong.

Scope (last ~year, per Vasu):
  - Shakkarpara: all complete batches Jun 2025 -> Jun 2026.
  - Rojmel: the one filled day (27 Jun 2026) + that month's opening stock.

Idempotent: clears the target date ranges first, so it's safe to re-run.

Run from the backend venv:
  backend/.venv/Scripts/python.exe scripts/import_history.py
"""

import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "backend"))

from app.db import SessionLocal, init_db  # noqa: E402
from app.modules.shakkarpara import engine  # noqa: E402
from app.modules.shakkarpara.models import Batch, BatchIngredient, OilSit  # noqa: E402
from app.modules.rojmel.models import (  # noqa: E402
    RojmelDay,
    RojmelExpenseLine,
    RojmelIncomeLine,
    RojmelSalesLine,
    RojmelStock,
)
from app.modules.rojmel.defaults import DEFAULT_PRODUCTS  # noqa: E402

SHAK_XLSX = ROOT / "source_excel" / "shakkarpara_filled.xlsx"
ROJMEL_XLSX = ROOT / "source_excel" / "rojmed_filled.xlsx"

SHAK_SHEETS = ["1.6.25", "1.7.25", "1.8.25", "1.9.25", "1.10.25", "1.11.25",
               "1.1.26", "1.2.26", "1.3.26", "1.4.26", "1.5.26", "1.6.26"]

_DATE_RE = re.compile(r"^\d{1,2}\.\d{1,2}\.\d{2,4}$")

INGREDIENT_CANON = {
    "oil": "Oil", "menda": "Menda", "elaichi": "Elaichi", "suger": "Sugar", "sugar": "Sugar",
    "ghee": "Ghee", "pelet": "Pelet", "box&plastic": "Box & Plastic", "masala": "Masala",
    "oil vaprayel": "Oil Vaprayel", "mansho": "Mansho",
}


def _num(v) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and _DATE_RE.match(v.strip()):
        for fmt in ("%d.%m.%y", "%d.%m.%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _canon_ingredient(raw: str) -> str:
    s = raw.strip().lower().replace(" ret", "").strip()
    return INGREDIENT_CANON.get(s, raw.strip().title())


def _find_blocks(ws):
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.strip().lower() == "date":
                d = _parse_date(ws.cell(c.row, c.column + 1).value)
                if d is not None:
                    yield c.row, c.column, d


def _extra_from_formula(wsf, r, bc) -> float:
    """Read the padtar cell formula (R+14, base+5) and return the +N constant, else 0."""
    f = wsf.cell(r + 14, bc + 5).value
    if isinstance(f, str):
        m = re.search(r"\)\s*\+\s*(\d+(?:\.\d+)?)\s*$", f.replace(" ", ""))
        if m:
            return float(m.group(1))
    return 0.0


def import_shakkarpara(db) -> None:
    wbv = openpyxl.load_workbook(SHAK_XLSX, data_only=True)
    wbf = openpyxl.load_workbook(SHAK_XLSX, data_only=False)

    imported = skipped_incomplete = failed = 0
    fails = []

    # idempotent: clear Jun 2025 .. Jun 2026 range
    db.query(Batch).filter(Batch.date >= date(2025, 6, 1), Batch.date <= date(2026, 6, 30)).delete()
    db.commit()

    for name in SHAK_SHEETS:
        if name not in wbv.sheetnames:
            print(f"  ! sheet {name} not found, skipping")
            continue
        wsv, wsf = wbv[name], wbf[name]
        for r, bc, d in _find_blocks(wsv):
            ings, prod, exp_total, exp_padtar, oil, exp_net = _extract(wsv, r, bc)
            if not prod or prod <= 0:
                skipped_incomplete += 1
                continue
            extra = _extra_from_formula(wsf, r, bc)

            # verify with the engine vs Excel cached values
            eng_lines = [engine.IngredientLine(name=i["name"], rate=i["rate"], usage=i["usage"],
                                               unit=i["unit"], is_oil_vaprayel=i["is_oil_vaprayel"]) for i in ings]
            eng_oil = engine.OilSit(**oil)
            res = engine.compute_batch(eng_lines, eng_oil, prod, extra)
            ok_total = res.total == exp_total
            ok_padtar = exp_padtar is None or res.padtar == exp_padtar
            if not (ok_total and ok_padtar):
                failed += 1
                fails.append((name, d.isoformat(), f"total {res.total} vs {exp_total}, padtar {res.padtar} vs {exp_padtar}"))
                continue

            batch = Batch(date=d, production_qty=prod, production_unit="kg", extra_per_unit=extra, notes=None)
            for idx, i in enumerate(ings):
                batch.ingredients.append(BatchIngredient(
                    name=i["name"], rate=i["rate"], usage=i["usage"], unit=i["unit"],
                    sort_order=idx, is_oil_vaprayel=i["is_oil_vaprayel"]))
            batch.oil_sit = OilSit(**oil)
            db.add(batch)
            imported += 1

    db.commit()
    print(f"\nShakkarpara: imported {imported}, skipped incomplete {skipped_incomplete}, verify-failed {failed}")
    for f in fails:
        print(f"   VERIFY FAIL {f[0]} {f[1]}: {f[2]}")


def _extract(ws, r, bc):
    ings = []
    for off in range(2, 12):
        rr = r + off
        raw = ws.cell(rr, bc).value
        if not isinstance(raw, str) or not raw.strip():
            continue
        unit = ws.cell(rr, bc + 3).value.strip() if isinstance(ws.cell(rr, bc + 3).value, str) else ""
        unit_l = unit.lower()
        # The oil-sit-linked row is the one measured in tins ("dabba") — identify by UNIT,
        # not by name. Newer sheets confusingly label the FIRST (batter-oil, unit "lot
        # bandhta") row "oil vaprayel" too; that row keeps its own typed usage.
        is_ov = unit_l == "dabba"
        if is_ov:
            name = "Oil Vaprayel"
        elif "vaprayel" in raw.lower():
            name = "Oil"  # mislabeled first row = the batter oil
        else:
            name = _canon_ingredient(raw)
        ings.append({
            "name": name,
            "rate": _num(ws.cell(rr, bc + 1).value),
            "usage": _num(ws.cell(rr, bc + 2).value),
            "unit": unit,
            "is_oil_vaprayel": is_ov,
            "expected_total": _num(ws.cell(rr, bc + 4).value),
        })
    prod = _num(ws.cell(r + 12, bc + 1).value)
    exp_total = _num(ws.cell(r + 12, bc + 4).value)
    praw = ws.cell(r + 14, bc + 5).value
    exp_padtar = float(praw) if isinstance(praw, (int, float)) else None
    oil = {
        "nava_dabba": _num(ws.cell(r + 17, bc + 1).value),
        "juna_dabba": _num(ws.cell(r + 17, bc + 2).value),
        "toppa": _num(ws.cell(r + 17, bc + 3).value),
        "parat_malela": _num(ws.cell(r + 17, bc + 4).value),
    }
    nraw = ws.cell(r + 17, bc + 5).value
    exp_net = float(nraw) if isinstance(nraw, (int, float)) else None
    return ings, prod, exp_total, exp_padtar, oil, exp_net


def import_rojmel(db) -> None:
    wb = openpyxl.load_workbook(ROJMEL_XLSX, data_only=True)
    ws = wb["Sheet1"]
    day_date = date(2026, 6, 27)

    # idempotent
    db.query(RojmelDay).filter(RojmelDay.date == day_date).delete()
    db.query(RojmelStock).filter(RojmelStock.year == 2026, RojmelStock.month == 6).delete()
    db.commit()

    canon_products = [p["name"] for p in DEFAULT_PRODUCTS]  # positional canonical names

    # sales: per-product totals in row 25 (cols B..K = 2..11), rate in row 2
    sales = []
    for col in range(2, 12):
        rate = _num(ws.cell(2, col).value)
        qty = _num(ws.cell(25, col).value)
        sales.append({"product": canon_products[col - 2], "rate": rate, "qty": qty})

    # income (cols M/N/O = 13/14/15), skip the factory-sales auto row (M20)
    income = []
    for row in range(3, 20):
        amt = ws.cell(row, 13).value
        desc = ws.cell(row, 14).value
        if isinstance(amt, (int, float)) and amt and str(desc or "").strip().lower() != "fatory sales":
            income.append({"description": str(desc or "").strip(), "amount": float(amt),
                           "note": str(ws.cell(row, 15).value or "").strip()})

    # expense (cols P/Q/R = 16/17/18), skip the totals rows (P20..P22)
    expense = []
    for row in range(3, 20):
        amt = ws.cell(row, 16).value
        desc = str(ws.cell(row, 17).value or "").strip()
        if isinstance(amt, (int, float)) and amt and desc.lower() not in ("totalkharcho", "cash o hand"):
            expense.append({"description": desc, "amount": float(amt),
                            "note": str(ws.cell(row, 18).value or "").strip()})

    # the extra carry-forward / pending block (rows 28-30, M/N + P/Q) -> preserve as notes
    note_bits = []
    for row in range(28, 32):
        for acol, dcol in ((13, 14), (16, 17)):
            amt = ws.cell(row, acol).value
            desc = str(ws.cell(row, dcol).value or "").strip()
            if isinstance(amt, (int, float)) and amt and desc:
                note_bits.append(f"{desc}: {amt:g}")
    notes = "Carried-over notes from sheet — " + "; ".join(note_bits) if note_bits else None

    day = RojmelDay(date=day_date, notes=notes)
    for idx, s in enumerate(sales):
        day.sales_lines.append(RojmelSalesLine(product=s["product"], rate=s["rate"], qty=s["qty"], sort_order=idx))
    for idx, m in enumerate(income):
        day.income_lines.append(RojmelIncomeLine(description=m["description"], amount=m["amount"], note=m["note"], sort_order=idx))
    for idx, m in enumerate(expense):
        day.expense_lines.append(RojmelExpenseLine(description=m["description"], amount=m["amount"], note=m["note"], sort_order=idx))
    db.add(day)

    # verify cashbook math vs the sheet's own totals
    from app.modules.rojmel import engine as rengine
    res = rengine.compute_day(
        [rengine.SalesLine(product=s["product"], rate=s["rate"], qty=s["qty"]) for s in sales],
        [rengine.MoneyLine(description=m["description"], amount=m["amount"]) for m in income],
        [rengine.MoneyLine(description=m["description"], amount=m["amount"]) for m in expense],
    )
    exp_income = _num(ws.cell(22, 13).value)   # M22 total
    exp_expense = _num(ws.cell(20, 16).value)  # P20 totalkharcho
    exp_cash = _num(ws.cell(21, 16).value)     # P21 cash o hand
    print(f"\nRojmel 27 Jun 2026 verification:")
    print(f"   income  {res.total_income}  vs sheet {exp_income}  {'OK' if res.total_income==exp_income else 'MISMATCH'}")
    print(f"   expense {res.total_expense} vs sheet {exp_expense} {'OK' if res.total_expense==exp_expense else 'MISMATCH'}")
    print(f"   cash    {res.cash_on_hand}  vs sheet {exp_cash}   {'OK' if res.cash_on_hand==exp_cash else 'MISMATCH'}")

    # stock: opening pieces per product (rows 29-38, D=opening 4, C=rate 3);
    # rows 29..38 map to the 10 products in the same order as the canonical list.
    for row in range(29, 39):
        if not str(ws.cell(row, 2).value or "").strip():
            continue
        opening = _num(ws.cell(row, 4).value)
        rate = _num(ws.cell(row, 3).value)
        canon = canon_products[row - 29]
        db.add(RojmelStock(year=2026, month=6, product=canon, rate=rate, opening_pic=opening))

    db.commit()
    print(f"   imported day with {len(sales)} products, {len(income)} income, {len(expense)} expense lines + June stock")


def main():
    init_db()
    db = SessionLocal()
    try:
        import_shakkarpara(db)
        import_rojmel(db)
    finally:
        db.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
