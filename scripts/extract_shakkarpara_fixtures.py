"""One-off extraction: read every batch block from the real
source_excel/03_shakkarpara.xlsx (20 monthly sheets, ~9-10 stacked batch
blocks each) and dump them as a JSON fixture the pytest suite can assert
the formula engine against.

The block layout is identical across all 20 sheets, just column-shifted
(early sheets: label col B, later sheets: label col A) — confirmed by
direct inspection. Relative to a block's "date" row R and label column
base_col:

    R      : label='date'                          col=base_col
    R+2..11: 10 ingredient rows                     label=base_col, rate=+1, usage=+2, unit=+3, total=+4
    R+12   : 'production'                           label=base_col, qty=+1, total_cost=+4
    R+14   : padtar value                           col=base_col+5
    R+17   : oil-sit values (nava/juna/toppa/parat/net)  cols base_col+1..+5

We detect blocks generically by scanning for a cell whose value is the
string "date" immediately followed (same row, next column) by a real
date — rather than hardcoding row numbers — so this is robust to the
known column drift between early/late sheets.
"""

import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

# Most sheets store the batch date as a real Excel date, but several sheets
# (confirmed by inspection: 1.2.23, 1.5.23, 1.6.23, 1.8.23, 01.10.23, 1.4.24)
# store it as a plain "D.M.YY" text string instead.
_DATE_STRING_RE = re.compile(r"^\d{1,2}\.\d{1,2}\.\d{2,4}$")

SOURCE_XLSX = Path(__file__).resolve().parent.parent / "source_excel" / "03_shakkarpara.xlsx"
OUTPUT_JSON = Path(__file__).resolve().parent.parent / "backend" / "tests" / "fixtures" / "shakkarpara_extracted.json"

INGREDIENT_ROW_OFFSETS = range(2, 12)  # R+2 .. R+11, 10 rows


def _num(value) -> float:
    return float(value) if isinstance(value, (int, float)) else 0.0


def _parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and _DATE_STRING_RE.match(value.strip()):
        return datetime.strptime(value.strip(), "%d.%m.%y").date()
    return None


def _find_block_starts(ws):
    """Yield (row, base_col) for every 'date' marker followed by a parseable date."""
    for row in range(1, ws.max_row + 1):
        for col in (1, 2):
            label_cell = ws.cell(row=row, column=col).value
            if isinstance(label_cell, str) and label_cell.strip().lower() == "date":
                if _parse_date(ws.cell(row=row, column=col + 1).value) is not None:
                    yield row, col


def _extract_block(ws, row: int, base_col: int, sheet_name: str) -> dict:
    date_value = _parse_date(ws.cell(row=row, column=base_col + 1).value)

    ingredients = []
    for offset in INGREDIENT_ROW_OFFSETS:
        r = row + offset
        name = ws.cell(row=r, column=base_col).value
        if not isinstance(name, str) or not name.strip():
            continue
        ingredients.append(
            {
                "name": name.strip(),
                "rate": _num(ws.cell(row=r, column=base_col + 1).value),
                "usage": _num(ws.cell(row=r, column=base_col + 2).value),
                "unit": (ws.cell(row=r, column=base_col + 3).value or "").strip()
                if isinstance(ws.cell(row=r, column=base_col + 3).value, str)
                else "",
                "expected_total": _num(ws.cell(row=r, column=base_col + 4).value),
            }
        )

    production_row = row + 12
    production_qty = _num(ws.cell(row=production_row, column=base_col + 1).value)
    expected_batch_total = _num(ws.cell(row=production_row, column=base_col + 4).value)

    padtar_value_row = row + 14
    expected_padtar_cell = ws.cell(row=padtar_value_row, column=base_col + 5).value
    expected_padtar = float(expected_padtar_cell) if isinstance(expected_padtar_cell, (int, float)) else None

    oil_sit_row = row + 17
    oil_sit = {
        "nava_dabba": _num(ws.cell(row=oil_sit_row, column=base_col + 1).value),
        "juna_dabba": _num(ws.cell(row=oil_sit_row, column=base_col + 2).value),
        "toppa": _num(ws.cell(row=oil_sit_row, column=base_col + 3).value),
        "parat_malela": _num(ws.cell(row=oil_sit_row, column=base_col + 4).value),
    }
    net_cell = ws.cell(row=oil_sit_row, column=base_col + 5).value
    expected_net_vaprash = float(net_cell) if isinstance(net_cell, (int, float)) else None

    return {
        "sheet": sheet_name,
        "date": date_value.isoformat(),
        "ingredients": ingredients,
        "production_qty": production_qty,
        "oil_sit": oil_sit,
        "expected_net_vaprash": expected_net_vaprash,
        "expected_batch_total": expected_batch_total,
        "expected_padtar": expected_padtar,
    }


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    batches = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row, base_col in _find_block_starts(ws):
            batches.append(_extract_block(ws, row, base_col, sheet_name))

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(batches, indent=2))
    print(f"Extracted {len(batches)} batches from {len(wb.sheetnames)} sheets -> {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
